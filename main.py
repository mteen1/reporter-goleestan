import openpyxl
from collections import defaultdict


def load_subjects(filename, subject_code_row=3):

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    subjects = {}
    for row in ws.iter_rows(min_row=2):
        subject_code = row[subject_code_row].value
        subjects[subject_code] = {
            "name": row[4].value,
            "major_code": row[25].value,
            "term_required": row[9].value,
        }
    return subjects


def load_students(filename):

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    students = {}

    for row in ws.iter_rows(min_row=2):
        student_id = row[6].value
        students[student_id] = {
            "name": f"{row[8].value} {row[7].value}",  # Combine first and last name
            "major_code": str(f"{row[22].value}{row[24].value}{row[26].value}"),
            "major_name": row[27].value,
        }
    return students


def load_scores(filename):

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    scores = defaultdict(dict)
    for row in ws.iter_rows(min_row=2):
        student_id = row[6].value
        subject_code = row[10].value.split("_")[0]  # Assuming subject code before "_"
        score = row[14].value
        scores[student_id][subject_code] = score
    return scores


def cleanscore(score: str) -> int:
    if "/" in score:
        num, float = score.split("/")
        score = int(f"{num}.{float}")
        return score
    else:
        return int(score)


def analyze_progress(subjects, students, scores, passing_score=12):
    """
    Analyzes student progress based on passed subjects and required terms.

    Args:
      subjects (dict): Dictionary of subject data.
      students (dict): Dictionary of student data.
      scores (dict): Dictionary of student scores.

    Returns:
      dict: A dictionary where keys are student IDs and values are dictionaries
            containing student information and progress details for each subject.
    """
    student_progress = {}
    for student_id, student_data in students.items():
        student_progress[student_id] = {
            "name": student_data["name"],
            "major_code": student_data["major_code"],
            "major_name": student_data["major_name"],
            "progress": {},
        }

        # Filter subjects based on student major code
        filtered_subjects = {
            subject_code: subject_data
            for subject_code, subject_data in subjects.items()
            if str(subject_data["major_code"]) == str(student_data["major_code"])
        }

        for subject_code, subject_data in filtered_subjects.items():
            try:
                score = scores.get(student_id, {}).get(str(subject_code), "بدون نمره")
                score = str(score).replace("/", ".")
                score = float(score)
                passed = True if score >= passing_score else False
            except ValueError:
                score = "-"
                passed = False

            # Add information for all subjects, even with missing scores
            student_progress[student_id]["progress"][subject_code] = {
                "passed": passed,  # Assume missing score implies not passed
                "score": score,
                "term_required": subject_data["term_required"],
                "name": subject_data["name"],
                "subject_code": str(subject_code),
            }

    # # Check if any student_progress dictionaries are empty
    # empty_progress = [
    #     student_id
    #     for student_id, progress in student_progress.items()
    #     if not progress["progress"]
    # ]

    return student_progress


import streamlit as st
import pandas as pd  # Import pandas for DataFrame manipulation


def main():
    """
    Main function to load data, analyze progress, and create a Streamlit dashboard.
    """
    subjects_file = "Subjects.xlsx"
    students_file = "Students.xlsx"
    scores_file = "Scores.xlsx"

    subjects = load_subjects(subjects_file)
    students = load_students(students_file)
    scores = load_scores(scores_file)

    student_progress = analyze_progress(subjects, students, scores)

    # Streamlit App
    import streamlit.components.v1 as components

    st.set_page_config(
        page_title="Nakhchivan pnu",
        page_icon="🧊",
        layout="wide",
        initial_sidebar_state="expanded",
        menu_items={
            "Get Help": "https://moharami.vip",
            "Report a bug": "https://www.extremelycoolapp.com/bug",
            "About": "pnu is a great university!",
        },
    )
    st.write(
        """
    <style>
    @font-face {
        font-family: 'vazirmatn';
        src: url("static/vazirmatn/fonts/ttf/Vazirmatn-Regular.woff2") format('woff2');
        }
    html, body, [class*="css"] {
    font-family: 'vazirmatn' !important;
    direction: rtl;
    }
    </style>
    """,
        unsafe_allow_html=True,
    )

    st.title(" پیشرفت تحصیلی دانشجو")

    # Select Student (optional)
    selected_student = None
    if len(student_progress) > 1:
        selected_student = st.selectbox("شماره دانشجویی", list(student_progress.keys()))

    # Display Student Progress
    if selected_student:
        student_data = student_progress[selected_student]
        st.subheader(f"دانشجو: {student_data['name']}")
        st.write(f"کد رشته: {student_data['major_code']}")
        major_name = subjects.get(student_data["major_code"], {}).get(
            "name", "Major Not Found"
        )  # Handle missing major code
        st.write(f"نام رشته: {student_data['major_name']}")

        # Option 1: Rename columns before displaying (using Pandas)
        progress_df = pd.DataFrame(student_data["progress"].values())
        progress_df.rename(
            columns={
                "passed": "وضعیت",
                "term_required": "ترم",
                "name": "نام درس",
                "subject_code": "کد",
                "score": "نمره",
            },
            inplace=True,
        )
        st.dataframe(progress_df, width=1500, hide_index=True)

        # Option 2: Use Streamlit's columns parameter for custom names
        # progress_table = st.dataframe(student_data["progress"].values(), columns=["Subject Name", "Status", "Term Required"])

    else:
        st.write("Please select a student to view their progress.")


if __name__ == "__main__":
    main()
